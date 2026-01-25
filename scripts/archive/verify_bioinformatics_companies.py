#!/usr/bin/env python3
"""Filter a company list to bioinformatics/biotech firms using web sources."""

from __future__ import annotations

import argparse
import csv
import re
import xml.etree.ElementTree as ET

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pathlib import Path
from io import BytesIO
from typing import Iterable
from urllib.parse import urlencode

import requests

USER_AGENT = "bioinfo-job-tracker/0.2"

ISHARES_IBB_HOLDINGS_URL = (
    "https://www.ishares.com/us/products/239699/ishares-biotechnology-etf/"
    "1467271812596.ajax?fileType=csv&fileName=IBB_holdings&dataType=fund"
)

WIKIDATA_SPARQL_ENDPOINT = "https://query.wikidata.org/sparql"
SSGA_XBI_HOLDINGS_URL = (
    "https://www.ssga.com/library-content/products/fund-data/etfs/us/"
    "holdings-daily-us-en-xbi.xlsx"
)
GLOBALX_GNOM_PAGE = "https://www.globalxetfs.com/funds/gnom/"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Filter companies to bioinformatics/biotech using web sources."
    )
    parser.add_argument(
        "--input",
        default="data/companies_all.csv",
        help="Input CSV with company_name header.",
    )
    parser.add_argument(
        "--output",
        default="data/companies.csv",
        help="Output CSV for filtered companies.",
    )
    parser.add_argument(
        "--output-unverified",
        default="data/companies_unverified.csv",
        help="Output CSV for companies not verified as bioinformatics/biotech.",
    )
    parser.add_argument(
        "--output-reference",
        default="data/biotech_reference_companies.csv",
        help="Output CSV with reference companies + source.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=20,
        help="HTTP timeout in seconds.",
    )
    return parser.parse_args()


def load_company_names(path: Path) -> list[str]:
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        names = [row["company_name"].strip() for row in reader if row.get("company_name")]
    return names


def normalize_name(value: str) -> str:
    value = value.lower()
    value = value.replace("&", "and")
    value = re.sub(r"\b(the|inc|incorporated|corp|corporation|co|company|llc|ltd|plc|gmbh|ag|sa|sarl|bv|kg|lp|llp)\b", "", value)
    value = re.sub(r"[^a-z0-9]+", "", value)
    return value


def read_allowlist(path: Path) -> set[str]:
    if not path.exists():
        return set()
    with path.open(encoding="utf-8") as handle:
        return {line.strip() for line in handle if line.strip() and not line.startswith("#")}


def fetch_ishares_ibb_holdings(timeout: int) -> list[str]:
    response = requests.get(
        ISHARES_IBB_HOLDINGS_URL,
        timeout=timeout,
        headers={"User-Agent": USER_AGENT},
    )
    response.raise_for_status()

    lines = response.text.splitlines()
    start_idx = None
    for idx, line in enumerate(lines):
        if line.startswith("Ticker,"):
            start_idx = idx
            break
    if start_idx is None:
        raise RuntimeError("Unable to find holdings header in IBB CSV.")

    reader = csv.DictReader(lines[start_idx:])
    names = []
    for row in reader:
        name = (row.get("Name") or "").strip()
        if name:
            names.append(name)
    return names


def fetch_wikidata_companies(timeout: int) -> list[str]:
    query = """
    SELECT DISTINCT ?companyLabel WHERE {
      ?company wdt:P31/wdt:P279* wd:Q90298876.
      SERVICE wikibase:label { bd:serviceParam wikibase:language "en". }
    }
    """
    params = {"query": query}
    url = f"{WIKIDATA_SPARQL_ENDPOINT}?{urlencode(params)}"
    response = requests.get(
        url,
        timeout=timeout,
        headers={
            "User-Agent": USER_AGENT,
            "Accept": "application/sparql-results+xml",
        },
    )
    response.raise_for_status()

    content_type = response.headers.get("content-type", "")
    if "xml" not in content_type:
        raise RuntimeError(f"Unexpected Wikidata response type: {content_type}")

    names = []
    root = ET.fromstring(response.text)
    ns = {"s": "http://www.w3.org/2005/sparql-results#"}
    for result in root.findall(".//s:result", ns):
        binding = result.find("s:binding[@name='companyLabel']/s:literal", ns)
        if binding is not None and binding.text:
            names.append(binding.text.strip())

    return names


def fetch_ssga_xbi_holdings(timeout: int) -> list[str]:
    response = requests.get(
        SSGA_XBI_HOLDINGS_URL,
        timeout=timeout,
        headers={"User-Agent": USER_AGENT},
    )
    response.raise_for_status()

    workbook = load_workbook(filename=BytesIO(response.content), read_only=True)
    sheet = workbook.active

    header_row = None
    for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
        if not row:
            continue
        if any(cell == "Name" for cell in row):
            header_row = row
            break
        if any(cell == "Security" for cell in row):
            header_row = row
            break

    if not header_row:
        raise RuntimeError("Unable to find header row in XBI holdings.")

    header = [str(cell).strip() if cell is not None else "" for cell in header_row]
    if "Name" in header:
        name_idx = header.index("Name")
    else:
        name_idx = header.index("Security")

    names = []
    data_started = False
    for row in sheet.iter_rows(values_only=True):
        if not data_started:
            if row == header_row:
                data_started = True
            continue
        if not row or name_idx >= len(row):
            continue
        name = row[name_idx]
        if not name or not isinstance(name, str):
            continue
        names.append(name.strip())

    return names


def _find_first_csv_link(html: str) -> str | None:
    soup = BeautifulSoup(html, "html.parser")
    for link in soup.find_all("a"):
        href = link.get("href") or ""
        if "holdings" in href and href.endswith(".csv"):
            return href
    return None


def fetch_globalx_gnom_holdings(timeout: int) -> list[str]:
    page_resp = requests.get(
        GLOBALX_GNOM_PAGE, timeout=timeout, headers={"User-Agent": USER_AGENT}
    )
    page_resp.raise_for_status()

    csv_link = _find_first_csv_link(page_resp.text)
    if not csv_link:
        raise RuntimeError("Unable to find GNOM holdings CSV link.")

    if csv_link.startswith("/"):
        csv_link = f"https://www.globalxetfs.com{csv_link}"

    csv_resp = requests.get(csv_link, timeout=timeout, headers={"User-Agent": USER_AGENT})
    csv_resp.raise_for_status()

    lines = csv_resp.text.splitlines()
    reader = csv.DictReader(lines)
    names = []
    for row in reader:
        name = (row.get("Name") or row.get("Security") or "").strip()
        if name:
            names.append(name)
    return names


def write_company_csv(path: Path, names: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["company_name"])
        for name in names:
            writer.writerow([name])


def write_reference_csv(path: Path, rows: Iterable[tuple[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["company_name", "source"])
        for name, source in rows:
            writer.writerow([name, source])


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)

    companies = load_company_names(input_path)

    ibb_names = fetch_ishares_ibb_holdings(args.timeout)
    xbi_names = fetch_ssga_xbi_holdings(args.timeout)
    gnom_names = fetch_globalx_gnom_holdings(args.timeout)
    wikidata_names = fetch_wikidata_companies(args.timeout)

    allowlist = read_allowlist(Path("data/bioinformatics_allowlist.txt"))
    denylist = read_allowlist(Path("data/bioinformatics_denylist.txt"))

    reference_rows = []
    reference_names = set()

    for name in ibb_names:
        reference_rows.append((name, "ishares_ibb_holdings"))
        reference_names.add(name)

    for name in xbi_names:
        reference_rows.append((name, "ssga_xbi_holdings"))
        reference_names.add(name)

    for name in gnom_names:
        reference_rows.append((name, "globalx_gnom_holdings"))
        reference_names.add(name)

    for name in wikidata_names:
        reference_rows.append((name, "wikidata_industry"))
        reference_names.add(name)

    for name in allowlist:
        reference_rows.append((name, "manual_allowlist"))
        reference_names.add(name)

    normalized_reference = {normalize_name(name) for name in reference_names}
    normalized_deny = {normalize_name(name) for name in denylist}

    verified = []
    unverified = []

    for name in companies:
        normalized = normalize_name(name)
        if normalized in normalized_deny:
            unverified.append(name)
            continue
        if normalized in normalized_reference:
            verified.append(name)
        else:
            unverified.append(name)

    write_company_csv(Path(args.output), verified)
    write_company_csv(Path(args.output_unverified), unverified)
    write_reference_csv(Path(args.output_reference), sorted(reference_rows))

    print(f"Verified: {len(verified)}")
    print(f"Unverified: {len(unverified)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
