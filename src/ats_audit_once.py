#!/usr/bin/env python3

import os
import sys
import json
import time
import signal
import requests
from datetime import datetime
from urllib.parse import urlparse
from openpyxl import load_workbook

# -------------------------
# CONFIG
# -------------------------
REQUEST_TIMEOUT = 10          # seconds per HTTP request
MAX_COMPANY_SECONDS = 20      # cap per company
GLOBAL_TIMEOUT_SECONDS = 600  # 10 minutes total

OUTPUT_PATH = "data/ats_audit_baseline.json"

TARGET_FILE_CANDIDATES = [
    "Bioinformatics_Job_Target_List.xlsx",
    "data/Bioinformatics_Job_Target_List.xlsx",
    "targets/Bioinformatics_Job_Target_List.xlsx",
]

HEADERS = {
    "User-Agent": "bioinfo-job-tracker/ats-audit"
}

# -------------------------
# TIMEOUT HANDLING
# -------------------------
def timeout_handler(signum, frame):
    raise TimeoutError("Global audit timeout reached")

signal.signal(signal.SIGALRM, timeout_handler)
signal.alarm(GLOBAL_TIMEOUT_SECONDS)

# -------------------------
# UTILITIES
# -------------------------
def find_targets_file():
    env_override = os.getenv("TARGETS_PATH")
    if env_override and os.path.exists(env_override):
        return env_override

    for path in TARGET_FILE_CANDIDATES:
        if os.path.exists(path):
            return path

    return None


def load_companies(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    companies = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        company, role, url = row[:3]
        if not company or not url:
            continue
        companies.append({
            "company": company.strip(),
            "careers_url": url.strip()
        })

    return companies


# -------------------------
# ATS DETECTION
# -------------------------
def detect_ats_from_url(url):
    host = urlparse(url).netloc.lower()

    if "myworkdayjobs.com" in host:
        return "workday"
    if "greenhouse.io" in host:
        return "greenhouse"
    if "lever.co" in host:
        return "lever"
    if "ashbyhq.com" in host:
        return "ashby"
    if "icims.com" in host:
        return "icims"

    return "unknown"


# -------------------------
# API ATTEMPTS
# -------------------------
def try_greenhouse(url):
    parts = urlparse(url).path.strip("/").split("/")
    if not parts:
        raise ValueError("Cannot extract Greenhouse org")

    org = parts[0]
    api = f"https://boards.greenhouse.io/v1/boards/{org}/jobs"
    r = requests.get(api, headers=HEADERS, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    data = r.json()
    return api, len(data.get("jobs", []))


def try_lever(url):
    org = urlparse(url).path.strip("/").split("/")[0]
    api = f"https://api.lever.co/v0/postings/{org}"
    r = requests.get(api, headers=HEADERS, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    data = r.json()
    return api, len(data)


def try_ashby(url):
    org = urlparse(url).path.strip("/").split("/")[0]
    api = f"https://jobs.ashbyhq.com/api/non-auth/jobs?organizationSlug={org}"
    r = requests.get(api, headers=HEADERS, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    data = r.json()
    return api, len(data.get("jobs", []))


def try_workday(url):
    parsed = urlparse(url)
    host = parsed.netloc
    parts = parsed.path.strip("/").split("/")

    # expect: /en-US/SITE
    if len(parts) < 2:
        raise ValueError("Cannot extract Workday site")

    site = parts[-1]
    tenant = host.split(".")[0]

    api = f"https://{host}/wday/cxs/{tenant}/{site}/jobs"
    r = requests.get(api, headers=HEADERS, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    data = r.json()
    jobs = data.get("jobPostings", [])
    return api, len(jobs)


# -------------------------
# MAIN AUDIT
# -------------------------
def audit_company(entry):
    company = entry["company"]
    url = entry["careers_url"]

    start = time.time()
    result = {
        "careers_url": url,
        "detected_ats": None,
        "attempts": [],
        "final_status": "unknown"
    }

    ats = detect_ats_from_url(url)
    result["detected_ats"] = ats

    try:
        if ats == "greenhouse":
            api, count = try_greenhouse(url)
        elif ats == "lever":
            api, count = try_lever(url)
        elif ats == "ashby":
            api, count = try_ashby(url)
        elif ats == "workday":
            api, count = try_workday(url)
        else:
            result["final_status"] = "unknown_ats"
            return result

        result["attempts"].append({
            "ats": ats,
            "api_url": api,
            "http_status": 200,
            "ok": True,
            "jobs_count": count,
            "error": None
        })
        result["final_status"] = "ok"

    except Exception as e:
        result["attempts"].append({
            "ats": ats,
            "api_url": None,
            "http_status": None,
            "ok": False,
            "jobs_count": None,
            "error": str(e)
        })
        result["final_status"] = "error"

    elapsed = time.time() - start
    if elapsed > MAX_COMPANY_SECONDS:
        result["final_status"] = "timeout"

    return result


def main():
    targets_path = find_targets_file()
    print(f"[AUDIT] targets_file={targets_path}")

    if not targets_path:
        print("[AUDIT] ERROR: targets file not found")
        sys.exit(1)

    companies = load_companies(targets_path)
    print(f"[AUDIT] loaded_companies={len(companies)}")

    if len(companies) == 0:
        raise RuntimeError("Loaded 0 companies. Check Excel content.")

    audit = {
        "metadata": {
            "run_utc": datetime.utcnow().isoformat(),
            "companies_scanned": len(companies)
        },
        "results": {}
    }

    for entry in companies:
        name = entry["company"]
        audit["results"][name] = audit_company(entry)

    os.makedirs("data", exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(audit, f, indent=2)

    print(f"[AUDIT] wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
